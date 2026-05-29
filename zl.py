from openpyxl import load_workbook
import pandas as pd
from config import health_report_path, strategy_excel_path  # 导入配置文件中的路径

def zl():
    down_excel = strategy_excel_path  # 使用配置文件中的路径
    down_sheetname = 'Sheet1'
    totle_excel = health_report_path  # 使用配置文件中的路径
    totle_sheetname = '战略'
    add_new_sheet(totle_excel, totle_sheetname)
    copy_excel(totle_excel, totle_sheetname, down_excel, down_sheetname)

    # 读取健康度判别表的sheet1和sheet2
    健康度判别表_sheet1 = pd.read_excel(health_report_path, sheet_name='Sheet1')
    健康度判别表_sheet2 = pd.read_excel(health_report_path, sheet_name='战略')

    # 定义需要复制的列名
    columns_to_copy = [
        '健康度', '红黄原因',
        '需求状态', '实施阶段', '项目状态', '需求分析牵头科室', '中心',
        '需求实施牵头人', '需求申请人部门', '需求主牵系统', '计划上线日期', '实际上线日期'
    ]

    # 遍历sheet2中的每一行，根据需求编号查找健康度判别表的sheet1中的对应行，并复制数据
    for index, row in 健康度判别表_sheet2.iterrows():
        需求编号 = row['需求编号']
        匹配行 = 健康度判别表_sheet1[健康度判别表_sheet1['需求编号'] == 需求编号]

        if not 匹配行.empty:
            # 获取匹配行的数据
            匹配数据 = 匹配行.iloc[0]

            # 更新sheet2中的数据
            for col in columns_to_copy:
                健康度判别表_sheet2.at[index, col] = 匹配数据[col]

    # 将计划上线日期列转换为短日期格式
    健康度判别表_sheet2['计划上线日期'] = pd.to_datetime(健康度判别表_sheet2['计划上线日期']).dt.strftime(
        '%Y-%m-%d')

    # 将实际上线日期列转换为短日期格式（如果需要）
    健康度判别表_sheet2['实际上线日期'] = pd.to_datetime(健康度判别表_sheet2['实际上线日期']).dt.strftime(
        '%Y-%m-%d')

    # 将更新后的健康度判别表保存回原文件
    with pd.ExcelWriter(health_report_path, mode='a',
                        if_sheet_exists='replace') as writer:
        健康度判别表_sheet2.to_excel(writer, sheet_name='战略', index=False)


def add_new_sheet(excel_path, new_sheet_name):
    workbook = load_workbook(excel_path)
    if new_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=new_sheet_name)
    workbook.save(excel_path)


def copy_excel(totle_excel, totle_sheetname, down_excel, down_sheetname):
    down = load_workbook(down_excel)
    totle = load_workbook(totle_excel)
    totle_sheet = totle[totle_sheetname]
    down_sheet = down[down_sheetname]
    # 两个for循环遍历整个excel的单元格内容
    for i, row in enumerate(down_sheet.iter_rows()):
        for j, cell in enumerate(row):
            totle_sheet.cell(row=i + 1, column=j + 1, value=cell.value)
    totle.save(totle_excel)