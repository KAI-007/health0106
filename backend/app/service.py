import importlib
import io
import math
import sys
from contextlib import redirect_stderr, redirect_stdout
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[1]
LEGACY_DIR = BASE_DIR / "legacy"

if str(LEGACY_DIR) not in sys.path:
    sys.path.insert(0, str(LEGACY_DIR))

APPROVAL_SHEET_MAP = {
    "行办会": "hbh",
    "行领导": "hld",
    "总经理室": "zjls",
    "部门科经理": "bmkjl",
}

CHART_TITLES = {
    "行办会": "行办会审议立项项目-健康度",
    "行领导": "行领导签报立项项目-健康度",
    "总经理室": "总经理室签报立项项目-健康度",
    "部门科经理": "部门科经理签报立项项目",
}

CHART_CATEGORIES = [
    "本部直属",
    "公金应用研发中心",
    "管理支持中心",
    "集团基础应用研发中心",
    "零售应用研发中心",
    "同业应用研发中心",
    "数据中心",
]

YEAR_ROWS = ["2026年度", "2025年度"]
APPROVAL_ROWS = ["行办会", "行领导", "总经理室", "部门科经理"]
HEALTH_KEYS = ["红", "黄", "绿"]
PROJECT_STATUS_KEYS = ["实施中", "待投产", "已投产", "小计", "其中，超期投产"]


async def _save_upload(upload_file: UploadFile, target_path: Path) -> None:
    content = await upload_file.read()
    target_path.write_bytes(content)


def _load_legacy_modules():
    config = importlib.import_module("config")
    main_module = importlib.import_module("generate_health_report_0106")
    return config, main_module


def _empty_health_row():
    return {key: 0 for key in HEALTH_KEYS}


def _empty_project_row():
    return {key: 0 for key in PROJECT_STATUS_KEYS}


def _empty_launch_row():
    return {
        "按计划上线数": 0,
        "实际上线数": 0,
        "按时上线率": "0%",
    }


def _empty_chart_card(title: str):
    return {
        "title": title,
        "series": {
            "红": [0] * len(CHART_CATEGORIES),
            "黄": [0] * len(CHART_CATEGORIES),
            "绿": [0] * len(CHART_CATEGORIES),
        },
        "max": 4,
    }


def _safe_read_excel(path: Path, sheet_name=0):
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()


def _normalize_center(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def _coerce_date(series):
    return pd.to_datetime(series, errors="coerce")


def _calculate_health_summary(workbook_path: Path):
    data = {name: _empty_health_row() for name in APPROVAL_ROWS}

    for approval_name, sheet_name in APPROVAL_SHEET_MAP.items():
        df = _safe_read_excel(workbook_path, sheet_name=sheet_name)
        if df.empty or "健康度" not in df.columns:
            continue
        for health in HEALTH_KEYS:
            data[approval_name][health] = int((df["健康度"].astype(str) == health).sum())

    total = _empty_health_row()
    for row in data.values():
        for health in HEALTH_KEYS:
            total[health] += row[health]
    data["合计"] = total
    return data


def _calculate_strategy_summary(workbook_path: Path):
    result = {year: {"战略数": 10, "实施中需求数": 0, **_empty_health_row()} for year in YEAR_ROWS}
    df = _safe_read_excel(workbook_path, sheet_name="战略")
    if df.empty:
        return result

    year_col = next((col for col in df.columns if "年度" in str(col)), None)
    if year_col is None:
        return result

    status_col = "项目状态" if "项目状态" in df.columns else ("需求状态" if "需求状态" in df.columns else None)
    health_col = "健康度" if "健康度" in df.columns else None

    for year in YEAR_ROWS:
        year_df = df[df[year_col].astype(str).str.contains(year[:4], na=False)]
        if year_df.empty or not status_col:
            continue

        scoped_df = year_df[year_df[status_col].astype(str).isin(["实施中", "待投产"])]
        result[year]["实施中需求数"] = int(len(scoped_df))

        if health_col:
            for health in HEALTH_KEYS:
                result[year][health] = int((scoped_df[health_col].astype(str) == health).sum())

    return result


def _calculate_project_summary(workbook_path: Path):
    data = {name: _empty_project_row() for name in APPROVAL_ROWS}

    launched_df = _safe_read_excel(workbook_path, sheet_name="jhsx_1z")
    launched_counts = {name: 0 for name in APPROVAL_ROWS}
    if not launched_df.empty and "审批级别校准" in launched_df.columns:
        for approval_name in APPROVAL_ROWS:
            launched_counts[approval_name] = int((launched_df["审批级别校准"].astype(str) == approval_name).sum())

    for approval_name, sheet_name in APPROVAL_SHEET_MAP.items():
        df = _safe_read_excel(workbook_path, sheet_name=sheet_name)
        if df.empty:
            continue

        status_col = "项目状态" if "项目状态" in df.columns else None
        if status_col:
            data[approval_name]["实施中"] = int((df[status_col].astype(str) == "实施中").sum())
            data[approval_name]["待投产"] = int((df[status_col].astype(str) == "待投产").sum())
        data[approval_name]["已投产"] = launched_counts[approval_name]
        data[approval_name]["小计"] = (
            data[approval_name]["实施中"]
            + data[approval_name]["待投产"]
            + data[approval_name]["已投产"]
        )
        data[approval_name]["其中，超期投产"] = 0

    total = _empty_project_row()
    for row in data.values():
        for key in PROJECT_STATUS_KEYS:
            total[key] += row[key]
    data["合计"] = total
    return data


def _launch_week_label(now: datetime):
    chinese_months = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十", "十一", "十二"]
    current_monday = now - timedelta(days=now.weekday())
    first_day = now.replace(day=1)
    first_monday_in_month = first_day + timedelta(days=(7 - first_day.weekday()) % 7)
    if current_monday.date() < first_monday_in_month.date():
        week_index = 1
    else:
        week_index = ((current_monday.date() - first_monday_in_month.date()).days // 7) + 1
    prev_week_index = max(1, week_index - 2)
    current_display_index = max(1, week_index - 1)
    month_text = f"{chinese_months[now.month - 1]}月"
    return f"{str(now.year)[-2:]}年{month_text}第{prev_week_index}、{current_display_index}周"


def _calculate_launch_summary(workbook_path: Path):
    data = {name: _empty_launch_row() for name in APPROVAL_ROWS}
    launch_title = f"按时上线率（{_launch_week_label(datetime.now())}）"

    df = _safe_read_excel(workbook_path, sheet_name="jhsx_2z")
    if not df.empty and "审批级别校准" in df.columns and "实际上线日期" in df.columns and "计划上线日期" in df.columns:
        actual_dates = _coerce_date(df["实际上线日期"])
        plan_dates = _coerce_date(df["计划上线日期"])
        actual_mask = actual_dates.notna()
        planned_mask = actual_mask & plan_dates.notna() & ((plan_dates - actual_dates).dt.days >= 0)

        for approval_name in APPROVAL_ROWS:
            approval_mask = df["审批级别校准"].astype(str) == approval_name
            actual_count = int((approval_mask & actual_mask).sum())
            planned_count = int((approval_mask & planned_mask).sum())
            ratio = f"{round((planned_count / actual_count) * 100)}%" if actual_count else "0%"

            data[approval_name]["按计划上线数"] = planned_count
            data[approval_name]["实际上线数"] = actual_count
            data[approval_name]["按时上线率"] = ratio

    total_planned = sum(data[row]["按计划上线数"] for row in APPROVAL_ROWS)
    total_actual = sum(data[row]["实际上线数"] for row in APPROVAL_ROWS)
    total_ratio = f"{round((total_planned / total_actual) * 100)}%" if total_actual else "0%"
    data["合计"] = {
        "按计划上线数": total_planned,
        "实际上线数": total_actual,
        "按时上线率": total_ratio,
    }
    return {"title": launch_title, "rows": data}


def _calculate_chart_cards(workbook_path: Path):
    cards = []
    for approval_name in APPROVAL_ROWS:
        sheet_name = APPROVAL_SHEET_MAP[approval_name]
        title = CHART_TITLES[approval_name]
        card = _empty_chart_card(title)
        df = _safe_read_excel(workbook_path, sheet_name=sheet_name)
        if not df.empty and "中心" in df.columns and "健康度" in df.columns:
            for idx, center in enumerate(CHART_CATEGORIES):
                center_df = df[df["中心"].apply(_normalize_center) == center]
                for health in HEALTH_KEYS:
                    card["series"][health][idx] = int((center_df["健康度"].astype(str) == health).sum())
            totals = [
                card["series"]["红"][i] + card["series"]["黄"][i] + card["series"]["绿"][i]
                for i in range(len(CHART_CATEGORIES))
            ]
            max_total = max(totals) if totals else 0
            card["max"] = max(4, int(math.ceil(max_total / 5.0) * 5) if max_total > 4 else max_total or 4)
        cards.append(card)
    return cards


def build_email_payload(workbook_path: Path):
    current_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    return {
        "header": {
            "greeting": "各位领导、同事，您好！",
            "summary": f"{datetime.now().year}年{datetime.now().month}月金融科技部尚未上线开发建设类项目情况如下。",
            "meta": f"一、健康度情况（数据来源：科管平台，取数时间：{current_time}）",
        },
        "healthSummary": _calculate_health_summary(workbook_path),
        "strategySummary": _calculate_strategy_summary(workbook_path),
        "projectSummary": _calculate_project_summary(workbook_path),
        "launchSummary": _calculate_launch_summary(workbook_path),
        "chartCards": _calculate_chart_cards(workbook_path),
    }


async def generate_report(health_file: UploadFile, strategy_file: UploadFile):
    with TemporaryDirectory() as temp_dir:
        workspace = Path(temp_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"健康度判别表_{timestamp}.xlsx"
        input_path = workspace / (health_file.filename or "健康度源数据.xlsx")
        strategy_path = workspace / (strategy_file.filename or "战略.xlsx")
        output_path = workspace / output_filename

        await _save_upload(health_file, input_path)
        await _save_upload(strategy_file, strategy_path)

        log_buffer = io.StringIO()
        try:
            config, main_module = _load_legacy_modules()
            config.input_file = str(input_path)
            config.output_file = str(output_path)
            config.strategy_excel_path = str(strategy_path)
            config.health_report_path = str(output_path)
            with redirect_stdout(log_buffer), redirect_stderr(log_buffer):
                main_module.main()
        except Exception as exc:
            logs = log_buffer.getvalue()
            detail = f"生成报表失败：{exc}"
            if logs:
                detail = f"{detail}\n\n执行日志：\n{logs}"
            raise HTTPException(status_code=500, detail=detail) from exc

        if not output_path.exists():
            logs = log_buffer.getvalue()
            detail = "报表生成失败，未找到输出文件"
            if logs:
                detail = f"{detail}\n\n执行日志：\n{logs}"
            raise HTTPException(status_code=500, detail=detail)

        final_output_path = BASE_DIR / output_filename
        final_output_path.write_bytes(output_path.read_bytes())
        return {
            "downloadName": output_filename,
            "logs": log_buffer.getvalue(),
        }


def download_report(filename: str):
    file_path = BASE_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="下载文件不存在")
    return file_path


def _style_email_range(ws, start_row, end_row, start_col, end_col, fill=None, bold=False, center=True):
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border
            cell.font = Font(name="宋体", size=12, bold=bold)
            cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill


def _write_matrix_table(ws, start_row, start_col, title, intro, headers_top, headers_bottom, body_rows, merge_ranges):
    ws.cell(start_row, 1, title).font = Font(name="宋体", size=16, bold=True)
    ws.cell(start_row + 1, 1, intro).font = Font(name="宋体", size=12)
    ws.merge_cells(start_row=start_row + 3, start_column=start_col + 1, end_row=start_row + 3, end_column=start_col + len(headers_bottom) - 1)
    ws.cell(start_row + 3, start_col + 1, headers_top)
    for idx, header in enumerate(headers_bottom, start=start_col):
        ws.cell(start_row + 4, idx, header)
    current_row = start_row + 5
    for row in body_rows:
        for col_idx, value in enumerate(row, start=start_col):
            ws.cell(current_row, col_idx, value)
        current_row += 1
    for merge_start_col, merge_end_col, merge_start_row, merge_end_row, value in merge_ranges:
        ws.merge_cells(start_row=merge_start_row, start_column=merge_start_col, end_row=merge_end_row, end_column=merge_end_col)
        ws.cell(merge_start_row, merge_start_col, value)
    _style_email_range(ws, start_row + 3, current_row - 1, start_col, start_col + len(headers_bottom) - 1, fill=PatternFill("solid", fgColor="D9E2F3"), bold=False)
    _style_email_range(ws, start_row + 5, current_row - 1, start_col, start_col + len(headers_bottom) - 1)
    return current_row


def _add_health_charts(ws, start_row, payload):
    data_start_col = 20
    chart_positions = ["A{row}", "J{row}", "A{row}", "J{row}"]
    for idx, chart in enumerate(payload["chartCards"]):
        col_base = data_start_col + idx * 5
        ws.cell(start_row, col_base, "中心")
        ws.cell(start_row, col_base + 1, "红")
        ws.cell(start_row, col_base + 2, "黄")
        ws.cell(start_row, col_base + 3, "绿")
        ws.cell(start_row, col_base + 4, "总数")
        for row_offset, category in enumerate(CHART_CATEGORIES, start=1):
            red = chart["series"]["红"][row_offset - 1]
            yellow = chart["series"]["黄"][row_offset - 1]
            green = chart["series"]["绿"][row_offset - 1]
            ws.cell(start_row + row_offset, col_base, category)
            ws.cell(start_row + row_offset, col_base + 1, red)
            ws.cell(start_row + row_offset, col_base + 2, yellow)
            ws.cell(start_row + row_offset, col_base + 3, green)
            ws.cell(start_row + row_offset, col_base + 4, red + yellow + green)
        chart_obj = BarChart()
        chart_obj.type = "col"
        chart_obj.style = 10
        chart_obj.title = chart["title"]
        chart_obj.y_axis.title = "数量"
        chart_obj.height = 8
        chart_obj.width = 12
        chart_obj.legend = None
        data = Reference(ws, min_col=col_base + 4, min_row=start_row, max_row=start_row + len(CHART_CATEGORIES))
        cats = Reference(ws, min_col=col_base, min_row=start_row + 1, max_row=start_row + len(CHART_CATEGORIES))
        chart_obj.add_data(data, titles_from_data=True)
        chart_obj.set_categories(cats)
        chart_obj.varyColors = False
        chart_obj.series[0].graphicalProperties.solidFill = "00B050"
        target_row = start_row + (idx // 2) * 18
        target_col_anchor = "A" if idx % 2 == 0 else "J"
        ws.add_chart(chart_obj, f"{target_col_anchor}{target_row}")


def _build_email_sheet(workbook_path: Path):
    payload = build_email_payload(workbook_path)
    wb = load_workbook(workbook_path)
    if "邮件" in wb.sheetnames:
        wb.remove(wb["邮件"])
    ws = wb.create_sheet("邮件")

    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    row = 1
    ws.cell(row, 1, payload["header"]["greeting"]).font = Font(name="宋体", size=12)
    row += 1
    ws.cell(row, 1, payload["header"]["summary"]).font = Font(name="宋体", size=12)
    row += 1
    ws.cell(row, 1, payload["header"]["meta"]).font = Font(name="宋体", size=12)

    row += 2
    health_rows = [[name, payload["healthSummary"][name]["红"], payload["healthSummary"][name]["黄"], payload["healthSummary"][name]["绿"]] for name in APPROVAL_ROWS + ["合计"]]
    ws.cell(row, 1, "（一）立项项目健康度情况").font = Font(name="宋体", size=16, bold=True)
    ws.cell(row + 1, 1, "按审批级别统计本期立项项目健康度情况，分别展示红、黄、绿三类项目数量。").font = Font(name="宋体", size=12)
    ws.merge_cells(start_row=row + 3, start_column=2, end_row=row + 3, end_column=4)
    ws.cell(row + 3, 1, "审批级别")
    ws.cell(row + 3, 2, "本期健康度")
    for idx, header in enumerate(["红", "黄", "绿"], start=2):
        ws.cell(row + 4, idx, header)
    ws.merge_cells(start_row=row + 3, start_column=1, end_row=row + 4, end_column=1)
    data_row = row + 5
    for item in health_rows:
        for idx, value in enumerate(item, start=1):
            ws.cell(data_row, idx, value)
        data_row += 1
    _style_email_range(ws, row + 3, data_row - 1, 1, 4, fill=PatternFill("solid", fgColor="D9E2F3"))
    _style_email_range(ws, row + 5, data_row - 1, 1, 4)

    row = data_row + 2
    strategy_rows = [[name, payload["strategySummary"][name]["战略数"], payload["strategySummary"][name]["实施中需求数"], payload["strategySummary"][name]["红"], payload["strategySummary"][name]["黄"], payload["strategySummary"][name]["绿"]] for name in YEAR_ROWS]
    ws.cell(row, 1, "战略项目计划立项健康度情况").font = Font(name="宋体", size=16, bold=True)
    ws.cell(row + 1, 1, "按年度统计战略数量、实施中需求数及对应健康度，用于反映战略项目计划立项整体推进情况。").font = Font(name="宋体", size=12)
    ws.merge_cells(start_row=row + 3, start_column=4, end_row=row + 3, end_column=6)
    ws.cell(row + 3, 1, "年度")
    ws.cell(row + 3, 2, "战略数")
    ws.cell(row + 3, 3, "实施中需求数")
    ws.cell(row + 3, 4, "健康度")
    for idx, header in enumerate(["红", "黄", "绿"], start=4):
        ws.cell(row + 4, idx, header)
    for col in [1, 2, 3]:
        ws.merge_cells(start_row=row + 3, start_column=col, end_row=row + 4, end_column=col)
    data_row = row + 5
    for item in strategy_rows:
        for idx, value in enumerate(item, start=1):
            ws.cell(data_row, idx, value)
        data_row += 1
    _style_email_range(ws, row + 3, data_row - 1, 1, 6, fill=PatternFill("solid", fgColor="D9E2F3"))
    _style_email_range(ws, row + 5, data_row - 1, 1, 6)

    row = data_row + 2
    ws.cell(row, 1, "（二）各中心项目健康度分布").font = Font(name="宋体", size=16, bold=True)
    ws.cell(row + 1, 1, "按审批级别分别展示各中心立项项目健康度分布情况，图表按两行两列进行排布。").font = Font(name="宋体", size=12)
    _add_health_charts(ws, row + 3, payload)

    row += 40
    project_rows = [[name, payload["projectSummary"][name]["实施中"], payload["projectSummary"][name]["待投产"], payload["projectSummary"][name]["已投产"], payload["projectSummary"][name]["小计"], payload["projectSummary"][name]["其中，超期投产"]] for name in APPROVAL_ROWS + ["合计"]]
    ws.cell(row, 1, "（三）项目阶段情况").font = Font(name="宋体", size=16, bold=True)
    ws.cell(row + 1, 1, "按审批级别统计项目阶段分布情况，分别展示实施中、待投产、已投产、小计及其中超期投产数量。").font = Font(name="宋体", size=12)
    ws.merge_cells(start_row=row + 3, start_column=2, end_row=row + 3, end_column=6)
    ws.cell(row + 3, 1, "审批级别")
    ws.cell(row + 3, 2, "项目阶段")
    for idx, header in enumerate(["实施中", "待投产", "已投产", "小计", "其中，超期投产"], start=2):
        ws.cell(row + 4, idx, header)
    ws.merge_cells(start_row=row + 3, start_column=1, end_row=row + 4, end_column=1)
    data_row = row + 5
    for item in project_rows:
        for idx, value in enumerate(item, start=1):
            ws.cell(data_row, idx, value)
        data_row += 1
    _style_email_range(ws, row + 3, data_row - 1, 1, 6, fill=PatternFill("solid", fgColor="D9E2F3"))
    _style_email_range(ws, row + 5, data_row - 1, 1, 6)

    row = data_row + 2
    launch_rows = [[name, payload["launchSummary"]["rows"][name]["按计划上线数"], payload["launchSummary"]["rows"][name]["实际上线数"], payload["launchSummary"]["rows"][name]["按时上线率"]] for name in APPROVAL_ROWS + ["合计"]]
    ws.cell(row, 1, "（四）按时上线率情况").font = Font(name="宋体", size=16, bold=True)
    ws.cell(row + 1, 1, "按审批级别统计计划上线数、实际上线数及按时上线率，用于衡量各审批级别项目投产达成情况。").font = Font(name="宋体", size=12)
    ws.merge_cells(start_row=row + 3, start_column=2, end_row=row + 3, end_column=4)
    ws.cell(row + 3, 1, "审批级别")
    ws.cell(row + 3, 2, payload["launchSummary"]["title"])
    for idx, header in enumerate(["按计划上线数", "实际上线数", "按时上线率"], start=2):
        ws.cell(row + 4, idx, header)
    ws.merge_cells(start_row=row + 3, start_column=1, end_row=row + 4, end_column=1)
    data_row = row + 5
    for item in launch_rows:
        for idx, value in enumerate(item, start=1):
            ws.cell(data_row, idx, value)
        data_row += 1
    _style_email_range(ws, row + 3, data_row - 1, 1, 4, fill=PatternFill("solid", fgColor="D9E2F3"))
    _style_email_range(ws, row + 5, data_row - 1, 1, 4)

    ws.cell(data_row + 2, 1, "以上为本期健康度情况，请审阅。").font = Font(name="宋体", size=12)
    wb.save(workbook_path)


async def export_email_workbook(email_file: UploadFile):
    with TemporaryDirectory() as temp_dir:
        workspace = Path(temp_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_filename = f"邮件表格_{timestamp}.xlsx"
        email_path = workspace / (email_file.filename or "邮件表格.xlsx")
        output_path = workspace / export_filename

        await _save_upload(email_file, email_path)
        output_path.write_bytes(email_path.read_bytes())

        try:
            _build_email_sheet(output_path)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"导出邮件工作表失败：{exc}") from exc

        final_output_path = BASE_DIR / export_filename
        final_output_path.write_bytes(output_path.read_bytes())
        return {
            "downloadName": export_filename,
        }


async def parse_email_table(email_file: UploadFile):
    with TemporaryDirectory() as temp_dir:
        workspace = Path(temp_dir)
        email_path = workspace / (email_file.filename or "邮件表格.xlsx")
        await _save_upload(email_file, email_path)

        try:
            return build_email_payload(email_path)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"解析邮件表格失败：{exc}") from exc
