from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import config


HEADER_FILL = PatternFill(fill_type='solid', start_color='D9EAF7', end_color='D9EAF7')
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)
BODY_FONT = Font(name='宋体', size=10)
HEADER_FONT = Font(name='宋体', size=10, bold=True)


def _display_width(value):
    if value is None:
        return 0
    text = str(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


def apply_sheet_style(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = BODY_FONT

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header_value = column_cells[0].value
        if header_value == '需求名称':
            worksheet.column_dimensions[column_letter].width = 30
            continue

        max_length = 0
        for cell in column_cells:
            max_length = max(max_length, _display_width(cell.value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 8), 80)


def zl():
    down_excel = config.strategy_excel_path
    down_sheetname = 'Sheet1'
    totle_excel = config.health_report_path
    totle_sheetname = '战略'
    add_new_sheet(totle_excel, totle_sheetname)
    copy_excel(totle_excel, totle_sheetname, down_excel, down_sheetname)

    健康度判别表_sheet1 = pd.read_excel(config.health_report_path, sheet_name='Sheet1')
    健康度判别表_sheet2 = pd.read_excel(config.health_report_path, sheet_name='战略')

    columns_to_copy = [
        '健康度', '红黄原因',
        '需求状态', '实施阶段', '项目状态', '需求分析牵头科室', '中心',
        '需求实施牵头人', '需求申请人部门', '需求主牵系统', '计划上线日期', '实际上线日期'
    ]

    for index, row in 健康度判别表_sheet2.iterrows():
        需求编号 = row['需求编号']
        匹配行 = 健康度判别表_sheet1[健康度判别表_sheet1['需求编号'] == 需求编号]

        if not 匹配行.empty:
            匹配数据 = 匹配行.iloc[0]
            for col in columns_to_copy:
                健康度判别表_sheet2.at[index, col] = 匹配数据[col]

    健康度判别表_sheet2['计划上线日期'] = pd.to_datetime(健康度判别表_sheet2['计划上线日期']).dt.strftime('%Y-%m-%d')
    健康度判别表_sheet2['实际上线日期'] = pd.to_datetime(健康度判别表_sheet2['实际上线日期']).dt.strftime('%Y-%m-%d')

    with pd.ExcelWriter(config.health_report_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
        健康度判别表_sheet2.to_excel(writer, sheet_name='战略', index=False)
        apply_sheet_style(writer.sheets['战略'])


def add_new_sheet(excel_path, new_sheet_name):
    workbook = load_workbook(excel_path)
    if new_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=new_sheet_name)
    workbook.save(excel_path)


def copy_excel(totle_excel, totle_sheetname, down_excel, down_sheetname):
    down = load_workbook(down_excel)
    totle = load_workbook(totle_excel)
    totle_sheet = totle[totle_sheetname]
    down_sheet = down[down_sheetname]
    for i, row in enumerate(down_sheet.iter_rows()):
        for j, cell in enumerate(row):
            totle_sheet.cell(row=i + 1, column=j + 1, value=cell.value)
    totle.save(totle_excel)
