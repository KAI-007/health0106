import os
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import zl
import zj
import hbh
import hld
import zjl
import bmkjl
import jhsx_2z
import config


HEADER_FILL = PatternFill(fill_type='solid', start_color='D9EAF7', end_color='D9EAF7')
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)
BODY_FONT = Font(name='宋体', size=10)
HEADER_FONT = Font(name='宋体', size=10, bold=True)


def _display_width(value):
    if value is None:
        return 0
    text = str(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


def apply_sheet_style_by_name(workbook_path, sheet_name):
    workbook = load_workbook(workbook_path)
    worksheet = workbook[sheet_name]

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = BODY_FONT

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header_value = column_cells[0].value
        if header_value == '需求名称':
            worksheet.column_dimensions[column_letter].width = 30
            continue

        max_length = 0
        for cell in column_cells:
            max_length = max(max_length, _display_width(cell.value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 8), 80)

    workbook.save(workbook_path)


def main():
    # 读取原始Excel文件
    global df, cell
    input_file = config.input_file
    output_file = config.output_file
    # 验证文件路径
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"文件未找到: {input_file}")

    # 尝试读取文件并捕获异常
    try:
        df = pd.read_excel(input_file)
        print("文件读取成功")
    except Exception as e:
        print(f"文件读取失败: {e}")
        return  # 如果文件读取失败，直接返回，避免后续处理

    print("原始列名:", df.columns.tolist())

    # 筛选需要的字段
    selected_columns = [
        '需求名称', '健康度', '红黄原因', '需求编号', '需求状态', '实施阶段', '项目状态',
        '需求分析牵头科室', '中心', '需求实施牵头人', '需求申请人部门', '需求主牵系统',

        '计划上线日期', '实际上线日期', '实际项目规模', '分析立项金额',
        '包含采购', 'b模式金额', '审批级别', '需求来源',
        '是否经过分管总会签', '是否经过行领导双签', '审批级别校准'
    ]

    # 清理列名中的多余空格
    df.columns = df.columns.str.strip()

    # 确保所有需要的列都在DataFrame中，不在的则添加空列
    missing_columns = [col for col in selected_columns if col not in df.columns]
    if missing_columns:
        print(f"以下列在文件中不存在，将添加为空列: {missing_columns}")
        for col in missing_columns:
            df[col] = None

    # 按照指定顺序排列列
    df = df[selected_columns]

    # 确保所有需要的列都在DataFrame中，不在的则添加空列
    for col in selected_columns:
        if col not in df.columns:
            df[col] = None

    # 按照指定顺序排列列
    df = df[selected_columns]

    def calculate_health_and_reason(row):
        demand_source = row['需求来源']  # 需求来源
        current_status = row['需求状态']
        plan_online_date = row['计划上线日期']
        actual_date = datetime.now().date()

        if current_status not in ["开发中", "开发完成", "已提交测试", "已通过测试", "已通过审核", "已计划上线"]:
            return '绿', ''

        # 如果需求来源为空或其他值，默认为绿色
        if pd.isnull(demand_source) or demand_source not in ['科管平台', '协同需求']:
            return '绿', ''

        # 对于科管平台和协同需求
        if demand_source in ['科管平台', '协同需求']:
            # 检查是否有计划上线日期
            if pd.notnull(plan_online_date):
                # 标准化日期格式
                try:
                    if isinstance(plan_online_date, datetime):
                        plan_date = plan_online_date.date()
                    elif isinstance(plan_online_date, str):
                        # 尝试将字符串转换为日期对象
                        plan_date = pd.to_datetime(plan_online_date).date()
                    else:
                        # 假设已经是date对象
                        plan_date = plan_online_date

                    # 判断当前日期是否晚于计划上线日期
                    is_overdue = actual_date > plan_date

                    if is_overdue:
                        if demand_source == '科管平台':
                            # 如果需求状态为已计划上线
                            if current_status == '已计划上线':
                                return '黄', '存在上线超期风险'
                            # 如果需求状态不为已计划上线
                            else:
                                return '红', '实施进度滞后，存在上线超期风险'
                        elif demand_source == '协同需求':
                            return '黄', '存在上线超期风险'
                except Exception:
                    # 如果日期转换失败，默认为绿色
                    pass

        # 默认返回绿色
        return '绿', ''

    # 应用规则计算健康度和红黄原因
    df[['健康度', '红黄原因']] = df.apply(calculate_health_and_reason, axis=1, result_type='expand')

    # 保存到新的Excel文件（在计算健康度和红黄原因之后）
    df.to_excel(output_file, index=False)

    # 定义实施阶段和项目状态的映射
    status_mapping = {
        '待启动': ('待实施', '待实施'),
        '分析结果确认中': ('待实施', '待实施'),
        '开发中': ('实施-项目开发', '实施中'),
        '开发完成': ('实施-项目开发', '实施中'),
        '排期及审核中': ('待实施', '待实施'),
        '需求创建中': ('准备', '准备'),
        '需求分析中': ('准备', '准备'),
        '需求提出中': ('准备', '准备'),
        '需求已退回': (None, None),
        '已取消': (None, None),
        '已计划上线': ('实施-项目投产', '待投产'),
        '已上线': ('投产', '投产'),
        '已提交测试': ('实施-测试', '实施中'),
        '已通过测试': ('实施-测试', '实施中'),
        '已通过审核': ('实施-测试', '实施中'),
        '已关闭': (None, None),
        '排队中': ('准备', '准备'),
        "首次上线": (None, '实施中'),
    }

    # 应用映射
    df['实施阶段'] = df['需求状态'].map(lambda x: status_mapping.get(x, (None, None))[0])
    df['项目状态'] = df['需求状态'].map(lambda x: status_mapping.get(x, (None, None))[1])

    # 定义中心的映射
    department_mapping = {
        '零售产品研发科': '零售应用研发中心',
        '零售渠道设计科': '零售应用研发中心',
        '零售渠道研发科': '零售应用研发中心',
        '零售经营研发科': '零售应用研发中心',
        '零售质控科': '零售应用研发中心',
        '公金产品研发科': '公金应用研发中心',
        '公金渠道设计科': '公金应用研发中心',
        '公金渠道研发科': '公金应用研发中心',
        '公金经营研发科': '公金应用研发中心',
        '贸金应用研发科': '公金应用研发中心',
        '公金质控科': '公金应用研发中心',
        '资金应用研发科': '同业应用研发中心',
        '托管应用研发科': '同业应用研发中心',
        '资管应用研发科': '同业应用研发中心',
        '投企管理科': '',
        '投企应用研发科': '集团基础应用研发中心',
        '集团办公研发科': '集团基础应用研发中心',
        '架构技术管理科': '',
        '测试管理科': '管理支持中心',
        '核心应用研发科': '集团基础应用研发中心',
        '支付清算研发科': '集团基础应用研发中心',
        '风控应用研发科': '集团基础应用研发中心',
        '数据应用研发科': '集团基础应用研发中心',
        '基础平台研发科': '管理支持中心',
        '应用支持科': '管理支持中心',
        '架构平台科': '管理支持中心',
        '安全管理科': '本部直属',
        '运维平台科': '数据中心'
    }

    # 定义审批级别校准的映射函数
    def calculate_approval_level_adjustment(row):
        # 获取相关字段值，处理可能的NaN值
        include_purchase = row.get('包含采购', 0) if pd.notnull(row.get('包含采购')) else 0
        signed_by_bank_leader = row.get('是否经过行领导双签', '') if pd.notnull(row.get('是否经过行领导双签')) else ''
        signed_by_deputy_general_manager = row.get('是否经过分管总会签', '') if pd.notnull(row.get('是否经过分管总会签')) else ''

        # 如果包含采购字段值为空，则按照0处理
        if pd.isnull(include_purchase):
            include_purchase = 0

        # 规则1: 若是否经过行领导双签为是，且包含采购大于300，则审批级别校准为行办会
        if signed_by_bank_leader == '是' and include_purchase > 300:
            return '行办会'

        # 规则2: 若是否经过行领导双签为是，且包含采购小于300或为空，则审批级别校准为行领导
        elif signed_by_bank_leader == '是' and (include_purchase < 300 or include_purchase == 0):
            return '行领导'

        # 规则3: 若是否经过行领导双签为否，且是否经过分管总双签为是，则审批级别校准为总经理室
        elif signed_by_bank_leader == '否' and signed_by_deputy_general_manager == '是':
            return '总经理室'

        # 规则4: 若是否经过行领导双签为否，且是否经过分管总会签为否，则审批级别校准为部门科经理
        elif signed_by_bank_leader == '否' and signed_by_deputy_general_manager == '否':
            return '部门科经理'

        # 默认返回空值
        else:
            return ''

    # 应用审批级别校准映射
    df['审批级别校准'] = df.apply(calculate_approval_level_adjustment, axis=1)

    # 预处理"需求分析牵头科室"字段
    df['需求分析牵头科室'] = df['需求分析牵头科室'].apply(
        lambda x: re.sub(r'^金融科技部研发测试中心\s*', '', str(x)) if pd.notnull(x) else x
    )
    # 预处理"需求分析牵头科室"字段
    df['需求分析牵头科室'] = df['需求分析牵头科室'].apply(
        lambda y: re.sub(r'^金融科技部数据中心\s*', '', str(y)) if pd.notnull(y) else y
    )
    df['需求分析牵头科室'] = df['需求分析牵头科室'].apply(
        lambda z: re.sub(r'^金融科技部\s*', '', str(z)) if pd.notnull(z) else z
    )
    # 应用映射
    df['中心'] = df['需求分析牵头科室'].map(department_mapping)

    # 保存到新的Excel文件
    df.to_excel(output_file, index=False)
    apply_sheet_style_by_name(output_file, 'Sheet1')
    #
    # # 处理单元格样式
    # from openpyxl import load_workbook
    # from openpyxl.styles import PatternFill
    #
    # # 加载工作簿
    # wb = load_workbook(output_file)
    # ws = wb.active
    #
    # # 定义颜色映射
    # color_mapping = {
    #     '绿': '00FF00',
    #     '黄': 'FFFF00',
    #     '红': 'FF0000'
    # }
    #
    # # 应用颜色
    # for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
    #     for cell in row:
    #         health_status = cell.value
    #         if health_status in color_mapping and health_status != '绿':
    #             cell.fill = PatternFill(start_color=color_mapping[health_status],
    #                                     end_color=color_mapping[health_status],
    #                                     fill_type='solid')
    #
    # # 保存修改后的工作簿
    # wb.save(output_file)
    print("判别表生成")
    zj.zj()
    print("在建完成")
    hbh.hbh()
    print("行办会完成")
    hld.hld()
    print("行领导完成")
    zjl.zjl()
    print("总经理室完成")
    bmkjl.bmkjl()
    print("部门科经理完成")
    jhsx_2z.jhsx_2z()
    print("计划上线周表完成")
    # b3.b3()
    # print("计划上线表1完成")
    # b4.b4()
    # print("计划上线表2完成")
    zl.zl()
    print("战略表完成")
    # mail_show.mail_to_csv()
    # print("邮件完成")


if __name__ == "__main__":
    main()