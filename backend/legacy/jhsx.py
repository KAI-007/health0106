import os
from datetime import datetime, timedelta

import pandas as pd
import config
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type='solid', start_color='D9EAF7', end_color='D9EAF7')
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)
BODY_FONT = Font(name='宋体', size=10)
HEADER_FONT = Font(name='宋体', size=10, bold=True)


def _display_width(value):
    if value is None:
        return 0
    text = str(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


def apply_sheet_style(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = BODY_FONT

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header_value = column_cells[0].value
        if header_value == '需求名称':
            worksheet.column_dimensions[column_letter].width = 30
            continue

        max_length = 0
        for cell in column_cells:
            max_length = max(max_length, _display_width(cell.value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 8), 80)


def jhsx():  # 将函数名从hbh改为jhsx，以匹配文件名
    # 读取健康度判别表
    input_file = config.output_file
    output_file = config.output_file

    # 验证文件路径
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"文件未找到: {input_file}")

    # 读取Excel文件中的"Sheet1"工作表
    try:
        df = pd.read_excel(input_file, sheet_name='Sheet1')
    except ValueError as e:
        # 如果找不到名为"Sheet1"的工作表，给出提示
        print(f"读取工作表时出错: {e}")
        return

    # 确保'实际上线日期'列存在且为日期格式
    if '实际上线日期' not in df.columns:
        print("错误：找不到'实际上线日期'列")
        return

    # 将'实际上线日期'列转换为日期格式
    df['实际上线日期'] = pd.to_datetime(df['实际上线日期'], errors='coerce')

    # 过滤掉无效的日期
    valid_dates_df = df.dropna(subset=['实际上线日期'])
    
    # 计算当前周的过去两周时间范围
    today = datetime.today()
    # 计算两周前的日期（14天前）
    start_date = today - timedelta(days=14)
    
    # 筛选实际上线日期在指定时间范围内的条目
    filtered_df = valid_dates_df[(valid_dates_df['实际上线日期'] >= start_date) & (valid_dates_df['实际上线日期'] <= today)]

    # 读取现有的Excel文件并保存到'计划上线表格'工作表
    with pd.ExcelWriter(output_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
        filtered_df.to_excel(writer, sheet_name='计划上线', index=False)
        apply_sheet_style(writer.sheets['计划上线'])


if __name__ == "__main__":
    jhsx()