import os

import pandas as pd
import config
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type='solid', start_color='D9EAF7', end_color='D9EAF7')
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)
BODY_FONT = Font(name='宋体', size=10)
HEADER_FONT = Font(name='宋体', size=10, bold=True)


def _display_width(value):
    if value is None:
        return 0
    text = str(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


def apply_sheet_style(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = BODY_FONT

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header_value = column_cells[0].value
        if header_value == '需求名称':
            worksheet.column_dimensions[column_letter].width = 30
            continue

        max_length = 0
        for cell in column_cells:
            max_length = max(max_length, _display_width(cell.value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 8), 80)


def bmkjl():
    input_file = config.output_file
    output_file = config.output_file

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"文件未找到: {input_file}")

    try:
        df = pd.read_excel(input_file, sheet_name='zj')
    except ValueError as e:
        print(f"读取工作表时出错: {e}")
        print("请确保已运行zj.py生成'在建'工作表")
        return

    filtered_df = df[df['审批级别校准'] == '部门科经理']

    with pd.ExcelWriter(output_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
        filtered_df.to_excel(writer, sheet_name='bmkjl', index=False)
        apply_sheet_style(writer.sheets['bmkjl'])
