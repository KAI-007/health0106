import os

import pandas as pd
import config
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type='solid', start_color='D9EAF7', end_color='D9EAF7')
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)
BODY_FONT = Font(name='宋体', size=10)
HEADER_FONT = Font(name='宋体', size=10, bold=True)


def _display_width(value):
    if value is None:
        return 0
    text = str(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


def apply_sheet_style(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = BODY_FONT

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header_value = column_cells[0].value
        if header_value == '需求名称':
            worksheet.column_dimensions[column_letter].width = 30
            continue

        max_length = 0
        for cell in column_cells:
            max_length = max(max_length, _display_width(cell.value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 8), 80)


def zj():
    # 读取健康度判别表
    input_file = config.output_file
    output_file = config.output_file

    # 验证文件路径
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"文件未找到: {input_file}")

    # 读取Excel文件
    df = pd.read_excel(input_file)

    # 筛选需求状态为"开发中", "开发完成", "已提交测试", "已通过测试", "已通过审核", "已计划上线"的需求条目
    filtered_df = df[
        df['需求状态'].isin(["开发中", "开发完成", "已提交测试", "已通过测试", "已通过审核", "已计划上线","首次上线"])]


    # 读取现有的Excel文件
    with pd.ExcelWriter(output_file, mode='a', if_sheet_exists='new', engine='openpyxl') as writer:
        filtered_df.to_excel(writer, sheet_name='zj', index=False)
        apply_sheet_style(writer.sheets['zj'])
